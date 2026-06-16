import openpyxl
from datetime import datetime
from functools import wraps
from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl.styles import Font, PatternFill
from django.db.models import Q
from users.forms import LoginForm
from users.models import CustomUser
from .form import EditEmprunteurForm, EditMaterielForm, EnregistrerEmprunteurForm, MaterielForm, EmpruntForm
from .models import Materiel,Classe,Tierce, Emprunt, LigneEmprunt,Operation
from .pdf_utils import build_simple_pdf
from .rules import materiel_a_un_emprunt_en_cours


def manager_required(view_func):
    @wraps(view_func)
    def _wrapped(request: HttpRequest, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("equipement:login")
        if request.user.type_user != "gestionnaire":
            return redirect("users:dashboard")
        return view_func(request, *args, **kwargs)

    return _wrapped



def login_view(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated:
        if request.user.type_user == CustomUser.TypeUser.ADMIN:
            return redirect(f"{reverse('users:dashboard')}?tab=dashboard")
        elif request.user.type_user == CustomUser.TypeUser.MANAGER:
                return redirect(f"{reverse('equipement:dashboard')}?tab=dashboard")
    form = LoginForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            email = form.cleaned_data["email"].lower()
            password = form.cleaned_data["password"]

            try:
                user = CustomUser.objects.get(email=email)
            except CustomUser.DoesNotExist:
                user = None

            if not user or (not user.is_active) or (not user.check_password(password)):
                messages.error(request, "Identifiants incorrects.")
            else:
                login(request, user)
            if request.user.type_user == CustomUser.TypeUser.ADMIN:
                return redirect(f"{reverse('users:dashboard')}?tab=dashboard")
            elif request.user.type_user == CustomUser.TypeUser.MANAGER:
                return redirect(f"{reverse('equipement:dashboard')}?tab=dashboard")
    return render(request, "equipement/login.html", {"form": form})


def logout_view(request: HttpRequest) -> HttpResponse:
    logout(request)
    messages.success(request, "Déconnexion effectuée.")
    return redirect("equipement:login")


def register_materiel(request: HttpRequest, form: MaterielForm) -> HttpResponse:
    if request.method != "POST":
        messages.error(request, "Action non autorisee.")
        return redirect(f"{reverse('equipement:dashboard')}?tab=equipement/ajouter")
    if form.is_valid():
        form.save()
        messages.success(request, "Equipement enregistre avec succes.")
        return redirect(f"{reverse('equipement:dashboard')}?tab=equipement/liste")
    messages.error(request, "Veuillez corriger les erreurs du formulaire equipement.")
    return redirect(f"{reverse('equipement:dashboard')}?tab=equipement/ajouter")


@login_required
def edit_equipement(request: HttpRequest, materiel_id: str) -> HttpResponse:
    materiel = get_object_or_404(Materiel, pk=materiel_id)

    if request.method == "POST":
        form = EditMaterielForm(request.POST, instance=materiel)
        if form.is_valid():
            form.save()
            messages.success(request, f"Equipement {materiel.nom} modifié avec succès.")
            return redirect(
                f"{reverse('equipement:dashboard')}?tab=detail&materiel_id={materiel_id}"
            )
        messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        form = EditMaterielForm(instance=materiel)

    return render(request, "equipement/edit_equipement.html", {
        "form": form,
        "materiel": materiel,
    })


def register_emprunteur(request: HttpRequest, form: EnregistrerEmprunteurForm) -> HttpResponse:
    if request.method != "POST":
        messages.error(request, "Action non autorisee.")
        return redirect(f"{reverse('equipement:dashboard')}?tab=emprunteur/ajouter")
    if form.is_valid():
        form.save()
        messages.success(request, "Emprunteur enregistre avec succes.")
        return redirect(f"{reverse('equipement:dashboard')}?tab=emprunteur/liste")
    messages.error(request, "Veuillez corriger les erreurs du formulaire emprunteur.")
    return redirect(f"{reverse('equipement:dashboard')}?tab=emprunteur/ajouter")


@login_required
def edit_emprunteur(request: HttpRequest, emprunteur_id: str) -> HttpResponse:
    emprunteur = get_object_or_404(Tierce, pk=emprunteur_id)

    if request.method == "POST":
        form = EditEmprunteurForm(request.POST, instance=emprunteur)
        if form.is_valid():
            form.save()
            messages.success(request, f"Emprunteur {emprunteur.prenom} {emprunteur.nom} modifié avec succès.")
            return redirect(
                f"{reverse('equipement:dashboard')}?tab=detail_emprunteur&emprunteur_id={emprunteur_id}"
            )
        messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        form = EditEmprunteurForm(instance=emprunteur)

    return render(request, "equipement/edit_emprunteur.html", {
        "form": form,
        "emprunteur": emprunteur,
    })


def dashboard(request: HttpRequest) -> HttpResponse:
    tab = request.GET.get("tab", "home")

    editFormMateriel = EditMaterielForm()
    editFormEmprunteur = EditEmprunteurForm()
    formMateriel = MaterielForm()
    formEmprunteur = EnregistrerEmprunteurForm()
    formEmprunt = EmpruntForm()

    action = request.POST.get("action")
    target = None

    materiel_detail = None
    emprunteur_detail = None

    emprunteurs = Tierce.objects.all()
    materiels_total = Materiel.objects.all()
    emprunts = Emprunt.objects.filter(Q(statut=Emprunt.Statut.APPROUVE) | Q(statut=Emprunt.Statut.REFUSE)).select_related(
        "materiels",
        "emprunteur",
    ).prefetch_related("lignes__materiel").order_by("-date_operation")
    classes = Classe.objects.all()
  
    categorie_filtre = request.GET.get("categorie", "")
    if categorie_filtre:
        materiels_total = materiels_total.filter(categorie=categorie_filtre)

    
    categories = Materiel.Categorie.choices

    materiels_recents = Materiel.objects.order_by("-id_materiel")[:5]
    emprunt_recents = Emprunt.objects.order_by("-date_operation")[:5]
    materiels_disponibles = [m for m in materiels_total if m.est_disponible()]
    materiels_en_pret = [m for m in materiels_total if m.est_en_pret()]
    materiels_en_maintenance = [m for m in materiels_total if m.est_en_maintenance()]
    materiels_hors_service = Materiel.objects.filter(etat="HORS SERVICE").order_by("nom", "id_materiel")
    emprunts_non_rendus = Emprunt.objects.filter(
        statut=Emprunt.Statut.APPROUVE,
        date_retour_reelle__isnull=True,
    ).select_related("emprunteur").prefetch_related("lignes__materiel").order_by("-date_operation")
    equipements_non_rendus_count = sum(emprunt.lignes.count() or 1 for emprunt in emprunts_non_rendus)
    if request.method == "POST":

        if action == "register_materiel":
            formMateriel = MaterielForm(request.POST)
            return register_materiel(request, formMateriel)

        if action == "register_emprunteur":
            formEmprunteur = EnregistrerEmprunteurForm(request.POST)
            return register_emprunteur(request, formEmprunteur)

    if tab == "detail" and request.GET.get("materiel_id"):
        materiel_detail = get_object_or_404(Materiel, pk=request.GET.get("materiel_id"))
    if tab == "detail_emprunteur" and request.GET.get("emprunteur_id"):
        emprunteur_detail = get_object_or_404(Tierce, pk=request.GET.get("emprunteur_id"))

    context = {
        "tab": tab,
        "target": target,
        "formMateriel": formMateriel,
        "formEmprunteur": formEmprunteur,
        "formEmprunt": formEmprunt,
        "classes":classes,
        "materiels_total": materiels_total,
        "materiels_recents": materiels_recents,
        "materiels_disponibles": materiels_disponibles,
        "materiels_en_pret": materiels_en_pret,
        "materiels_en_maintenance": materiels_en_maintenance,
        "materiels_hors_service": materiels_hors_service,
        "emprunts_non_rendus": emprunts_non_rendus,
        "equipements_non_rendus_count": equipements_non_rendus_count,
        "materiel_detail": materiel_detail,
        "editMaterielForm": editFormMateriel,
        "editEmprunteurForm": editFormEmprunteur,
        "emprunteur_detail": emprunteur_detail,
        "emprunteurs": emprunteurs,
        "emprunts": emprunts,
        "statuts_emprunt": Emprunt.Statut.choices,
        "emprunt_recents": emprunt_recents,
        "categories": categories,
        "categorie_filtre": categorie_filtre,
        "today_date": timezone.localdate(),
    }
    return render(request, "equipement/dashboard.html", context)


@login_required
def rapport_pdf(request: HttpRequest) -> HttpResponse:
    materiels_hors_service = Materiel.objects.filter(etat="HORS SERVICE").order_by("nom", "id_materiel")
    emprunts_non_rendus = Emprunt.objects.filter(
        statut=Emprunt.Statut.APPROUVE,
        date_retour_reelle__isnull=True,
    ).select_related("emprunteur").prefetch_related("lignes__materiel").order_by("-date_operation")

    generated_at = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")
    lines = [
        f"Genere le : {generated_at}",
        f"Nombre d'equipements hors service : {materiels_hors_service.count()}",
        f"Nombre d'equipements non rendus : {sum(emprunt.lignes.count() or 1 for emprunt in emprunts_non_rendus)}",
        "",
        "## Equipements hors service",
    ]

    if materiels_hors_service.exists():
        for materiel in materiels_hors_service:
            lines.append(
                f"- {materiel.id_materiel} | {materiel.nom} | {materiel.get_categorie_display()} | "
                f"Marque: {materiel.marque} | Serie: {materiel.numero_serie} | Couleur: {materiel.couleur}"
            )
    else:
        lines.append("Aucun equipement hors service.")

    lines.extend(["", "## Equipements non rendus"])
    if emprunts_non_rendus.exists():
        for emprunt in emprunts_non_rendus:
            emprunteur = emprunt.emprunteur
            materiels = list(emprunt.lignes.all())
            if materiels:
                for ligne in materiels:
                    materiel = ligne.materiel
                    lines.append(
                        f"- {materiel.id_materiel} | {materiel.nom} | Dernier emprunteur: "
                        f"{emprunteur.prenom} {emprunteur.nom} ({emprunteur.id_Tierce}) | "
                        f"Date emprunt: {emprunt.date_operation:%d/%m/%Y} | Retour prevu: "
                        f"{emprunt.date_retour_prevue or 'Non renseigne'}"
                    )
            else:
                materiel = emprunt.materiels
                lines.append(
                    f"- {materiel.id_materiel} | {materiel.nom} | Dernier emprunteur: "
                    f"{emprunteur.prenom} {emprunteur.nom} ({emprunteur.id_Tierce}) | "
                    f"Date emprunt: {emprunt.date_operation:%d/%m/%Y} | Retour prevu: "
                    f"{emprunt.date_retour_prevue or 'Non renseigne'}"
                )
    else:
        lines.append("Aucun equipement non rendu.")

    pdf = build_simple_pdf("Rapport des equipements", lines)
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="rapport_equipements.pdf"'
    return response


def detail_materiel(request, materiel_id):
    get_object_or_404(Materiel, id_materiel=materiel_id)
    return redirect(f"{reverse('equipement:dashboard')}?tab=detail&materiel_id={materiel_id}")


def retirer_materiel(request, materiel_id):
    if request.method != "POST":
        messages.error(request, "Action non autorisee.")
        return redirect(f"{reverse('equipement:dashboard')}?tab=equipement/liste")
    materiel = get_object_or_404(Materiel, id_materiel=materiel_id)
    nom = materiel.nom
    materiel.delete()
    messages.success(request, f"Equipement {nom} retire avec succes.")
    return redirect(f"{reverse('equipement:dashboard')}?tab=equipement/liste")




@login_required
def enregistrer_emprunt_view(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = EmpruntForm(request.POST)
        if form.is_valid():
            materiels_choisis = form.cleaned_data["materiels"]
            date_operation = form.cleaned_data["date_operation"]
            now = timezone.now()

            if timezone.is_naive(date_operation):
                date_operation = timezone.make_aware(date_operation)

            statut = Emprunt.Statut.APPROUVE if date_operation <= now else Emprunt.Statut.EN_ATTENTE

            if statut == Emprunt.Statut.APPROUVE:
                indisponibles = [m for m in materiels_choisis if not m.est_disponible()]
                if indisponibles:
                    noms = ", ".join(m.nom for m in indisponibles)
                    messages.error(request, f"Matériel(s) non disponible(s) : {noms}")
                    return redirect(f"{reverse('equipement:dashboard')}?tab=emprunts/creer")

            emprunt = Emprunt.objects.create(
                materiels=materiels_choisis[0],
                date_operation=date_operation,
                notes=form.cleaned_data["notes"],
                type_operation=Operation.TypeOperation.EMPRUNT,
                emprunteur=form.cleaned_data["emprunteur"],
                classe=form.cleaned_data["emprunteur"].get_classe(),
                date_retour_prevue=form.cleaned_data["date_retour_prevue"],
                statut=statut,

            )

            for materiel in materiels_choisis:
                LigneEmprunt.objects.create(emprunt=emprunt, materiel=materiel)
                if statut == Emprunt.Statut.APPROUVE:
                    materiel.mettre_en_pret()

            messages.success(
                request,
                f"Emprunt enregistré avec succès ({len(materiels_choisis)} équipement(s), statut : {emprunt.get_statut_display()})."
            )
            return redirect(f"{reverse('equipement:dashboard')}?tab=emprunts/liste")

        erreurs = []
        for field_errors in form.errors.values():
            erreurs.extend(field_errors)
        messages.error(request, " ".join(erreurs) or "Erreur dans le formulaire.")
        return redirect(f"{reverse('equipement:dashboard')}?tab=emprunts/creer")
    return redirect(f"{reverse('equipement:dashboard')}?tab=emprunts/creer")


def exporter_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materiels"

    headers = ['id_materiel', 'nom', 'couleur', 'categorie', 'etat', 'marque','numero_serie ']
    ws.append(headers)

    for m in Materiel.objects.all():
        ws.append([
            str(m.id_materiel),
            m.nom or "",
            m.couleur or "",
            m.categorie or "",
            m.etat or "",
            m.marque or "",
            m.numero_serie or "",
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=materiels.xlsx'
    wb.save(response)
    return response


def import_materiels(request):
    if request.method != "POST":
        return redirect(f"{reverse('equipement:dashboard')}?tab=import")

    fichier = request.FILES.get("file")

    if not fichier:
        messages.error(request, "Veuillez sélectionner un fichier.")
        return redirect(f"{reverse('equipement:dashboard')}?tab=import")

    if not fichier.name.endswith(".xlsx"):
        messages.error(request, "Fichier invalide (.xlsx requis).")
        return redirect(f"{reverse('equipement:dashboard')}?tab=import")

    try:
        wb = openpyxl.load_workbook(fichier)
        ws = wb.active
    except Exception as e:
        messages.error(request, f"Erreur lecture fichier : {e}")
        return redirect(f"{reverse('equipement:dashboard')}?tab=import")

    importes = 0
    ignores = 0
    etats_bloques = 0

    valeurs_valides = [c[0] for c in Materiel.ETAT_CHOICES]

    for row in ws.iter_rows(min_row=2, values_only=True):

        if not row:
            ignores += 1
            continue

        row = [str(cell).strip() if cell else "" for cell in row]

        if len(row) < 6:
            ignores += 1
            continue

        id_materiel, nom, couleur, categorie, etat, marque = row[:6]

        if not id_materiel or not nom:
            ignores += 1
            continue

        etat = etat.upper()
        if etat not in valeurs_valides:
            etat = "DISPONIBLE"

        defaults = {
            "nom": nom,
            "couleur": couleur,
            "categorie": categorie,
            "etat": etat,
            "marque": marque,
        }
        materiel_existant = Materiel.objects.filter(id_materiel=id_materiel).first()
        if (
            materiel_existant
            and materiel_existant.etat != etat
            and materiel_a_un_emprunt_en_cours(materiel_existant)
        ):
            defaults["etat"] = materiel_existant.etat
            etats_bloques += 1

        Materiel.objects.update_or_create(id_materiel=id_materiel, defaults=defaults)
        importes += 1

    messages.success(
        request,
        f"Import terminé : {importes} lignes traitées, {ignores} ignorées."
    )
    return redirect(f"{reverse('equipement:dashboard')}?tab=equipement/liste")


def detail_materiel(request, materiel_id):
    get_object_or_404(Materiel, id_materiel=materiel_id)
    return redirect(f"{reverse('equipement:dashboard')}?tab=detail&materiel_id={materiel_id}")


def detail_emprunteur(request, emprunteur_id):
    get_object_or_404(Tierce, id_Tierce=emprunteur_id)
    return redirect(f"{reverse('equipement:dashboard')}?tab=detail_emprunteur&emprunteur_id={emprunteur_id}")


def retirer_emprunteur(request: HttpRequest, emprunteur_id: str) -> HttpResponse:
    if request.method != "POST":
        messages.error(request, "Action non autorisee.")
        return redirect(f"{reverse('equipement:dashboard')}?tab=emprunteur/liste")

    emprunteur = get_object_or_404(Tierce, id_Tierce=emprunteur_id)
    nom_complet = f"{emprunteur.prenom} {emprunteur.nom}".strip()
    emprunteur.delete()
    messages.success(request, f"Emprunteur {nom_complet} supprime avec succes.")
    return redirect(f"{reverse('equipement:dashboard')}?tab=emprunteur/liste")

def import_emprunts(request):
    """
    Importe des emprunts depuis un fichier .xlsx.

    Colonnes attendues (dans l'ordre) :
        id | materiels | emprunteur | date_operation | date_retour_prevue | notes | statut

    Règles métier appliquées :
    - Si l'id correspond à un emprunt existant → mise à jour sans toucher à l'état des matériels.
    - Un nouvel emprunt APPROUVE exige que tous ses matériels soient DISPONIBLES.
    - Un emprunt RETOURNE doit avoir une date_retour_reelle ; si absente on utilise aujourd'hui.
    - Un emprunt REFUSE ou EN_ATTENTE ne modifie pas l'état des matériels.
    - La date de retour prévue ne peut pas être antérieure à la date d'opération.
    """
    TAB_RETOUR = "emprunts/liste"

    if request.method != "POST":
        return redirect(f"{reverse('equipement:dashboard')}?tab={TAB_RETOUR}")

    fichier = request.FILES.get("file")
    if not fichier:
        messages.error(request, "Veuillez sélectionner un fichier.")
        return redirect(f"{reverse('equipement:dashboard')}?tab={TAB_RETOUR}")
    if not fichier.name.endswith(".xlsx"):
        messages.error(request, "Fichier invalide (.xlsx requis).")
        return redirect(f"{reverse('equipement:dashboard')}?tab={TAB_RETOUR}")

    try:
        wb = openpyxl.load_workbook(fichier)
        ws = wb.active
    except Exception as e:
        messages.error(request, f"Erreur lecture fichier : {e}")
        return redirect(f"{reverse('equipement:dashboard')}?tab={TAB_RETOUR}")

    importes = 0
    mis_a_jour = 0
    ignores = 0
    erreurs = []

    statuts_valides = [c[0] for c in Emprunt.Statut.choices] 
    today = timezone.localdate()

    def _parse_date(valeur: str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(valeur, fmt).date()
            except ValueError:
                continue
        return None

    def _resoudre_emprunteur(nom_complet: str):
        parties = nom_complet.strip().split()
        obj = None
        if len(parties) >= 2:
            obj = (
                Tierce.objects.filter(prenom__iexact=parties[0], nom__iexact=" ".join(parties[1:])).first()
                or Tierce.objects.filter(nom__iexact=parties[0], prenom__iexact=" ".join(parties[1:])).first()
            )
        return obj or Tierce.objects.filter(nom__iexact=nom_complet).first()

    def _resoudre_materiels(noms_bruts: str, numero_ligne: int):
        objs, errs = [], []
        for token in [n.strip() for n in noms_bruts.split(",") if n.strip()]:
            try:
                objs.append(Materiel.objects.get(nom__iexact=token))
            except Materiel.DoesNotExist:
                try:
                    objs.append(Materiel.objects.get(id_materiel__iexact=token))
                except Materiel.DoesNotExist:
                    errs.append(f"Ligne {numero_ligne} : matériel '{token}' introuvable.")
            except Materiel.MultipleObjectsReturned:
                errs.append(f"Ligne {numero_ligne} : plusieurs matériels correspondent à '{token}', utilisez l'id_materiel.")
        return objs, errs

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        if not any(row):
            continue

        row = [str(cell).strip() if cell is not None else "" for cell in row]

        if len(row) < 5:
            erreurs.append(f"Ligne {i} : pas assez de colonnes ({len(row)}, minimum 5 attendues).")
            ignores += 1
            continue

        emprunt_id     = row[0]
        noms_materiels = row[1]
        nom_emprunteur = row[2]
        date_op_brut   = row[3]
        retour_brut    = row[4]
        notes          = row[5] if len(row) > 5 else ""
        statut_brut    = row[6].upper().strip() if len(row) > 6 and row[6] else ""

        if not noms_materiels:
            erreurs.append(f"Ligne {i} : colonne 'materiels' vide.")
            ignores += 1
            continue
        if not nom_emprunteur:
            erreurs.append(f"Ligne {i} : colonne 'emprunteur' vide.")
            ignores += 1
            continue
        if not retour_brut:
            erreurs.append(f"Ligne {i} : date de retour prévue manquante.")
            ignores += 1
            continue

        date_operation = _parse_date(date_op_brut) if date_op_brut else today
        retour_prevu_date = _parse_date(retour_brut)

        if retour_prevu_date is None:
            erreurs.append(f"Ligne {i} : format de date invalide '{retour_brut}'.")
            ignores += 1
            continue

        if date_operation and retour_prevu_date < date_operation:
            erreurs.append(f"Ligne {i} : la date de retour ({retour_prevu_date}) est antérieure à la date d'emprunt ({date_operation}).")
            ignores += 1
            continue

        statut = statut_brut if statut_brut in statuts_valides else Emprunt.Statut.APPROUVE

        emprunteur_obj = _resoudre_emprunteur(nom_emprunteur)
        if not emprunteur_obj:
            erreurs.append(f"Ligne {i} : emprunteur '{nom_emprunteur}' introuvable en base.")
            ignores += 1
            continue

        materiels_objs, errs_materiels = _resoudre_materiels(noms_materiels, i)
        if errs_materiels:
            erreurs.extend(errs_materiels)
            ignores += 1
            continue
        if not materiels_objs:
            erreurs.append(f"Ligne {i} : aucun matériel valide trouvé.")
            ignores += 1
            continue

        emprunt_id_int = int(emprunt_id) if emprunt_id.isdigit() else None
        emprunt_existant = Emprunt.objects.filter(id=emprunt_id_int).first() if emprunt_id_int else None

        try:
            if emprunt_existant:
                # ── MISE À JOUR ──────────────────────────────────────────────
                # On ne retouche PAS l'état des matériels : ils sont déjà
                # dans le bon état en base, l'import ne doit pas les perturber.
                ancien_statut = emprunt_existant.statut

                emprunt_existant.emprunteur = emprunteur_obj
                emprunt_existant.date_retour_prevue = retour_prevu_date
                emprunt_existant.notes = notes
                emprunt_existant.statut = statut

                # Règle : RETOURNE sans date réelle → on met aujourd'hui
                if statut == Emprunt.Statut.RETOURNE and not emprunt_existant.date_retour_reelle:
                    emprunt_existant.date_retour_reelle = today

                emprunt_existant.save()

                # Sync des lignes sans toucher à l'état des matériels
                ids_existants = set(emprunt_existant.lignes.values_list("materiel_id", flat=True))
                ids_nouveaux = {m.id_materiel for m in materiels_objs}

                # Supprimer uniquement les lignes qui disparaissent
                emprunt_existant.lignes.exclude(materiel_id__in=ids_nouveaux).delete()

                # Ajouter uniquement les nouvelles lignes
                for m in materiels_objs:
                    if m.id_materiel not in ids_existants:
                        LigneEmprunt.objects.create(emprunt=emprunt_existant, materiel=m)

                # Si le statut passe à RETOURNE, libérer les matériels EN PRET
                if statut == Emprunt.Statut.RETOURNE and ancien_statut != Emprunt.Statut.RETOURNE:
                    for m in materiels_objs:
                        if m.est_en_pret():
                            m.retourner()

                mis_a_jour += 1

            else:
                # ── CRÉATION ────────────────────────────────────────────────
                # Nouvel emprunt APPROUVE : tous les matériels doivent être disponibles
                if statut == Emprunt.Statut.APPROUVE:
                    indisponibles = [m for m in materiels_objs if not m.est_disponible()]
                    if indisponibles:
                        noms_ind = ", ".join(m.nom for m in indisponibles)
                        erreurs.append(f"Ligne {i} : matériel(s) non disponible(s) : {noms_ind}. Emprunt ignoré.")
                        ignores += 1
                        continue

                date_retour_reelle = today if statut == Emprunt.Statut.RETOURNE else None
                emprunt = Emprunt.objects.create(
                    materiels=materiels_objs[0],
                    emprunteur=emprunteur_obj,
                    date_retour_prevue=retour_prevu_date,
                    date_retour_reelle=date_retour_reelle,
                    statut=statut,
                    notes=notes,
                    type_operation=Operation.TypeOperation.EMPRUNT,
                )

                for m in materiels_objs:
                    LigneEmprunt.objects.create(emprunt=emprunt, materiel=m)

                # Mise à jour de l'état des matériels uniquement à la création
                if statut == Emprunt.Statut.APPROUVE:
                    for m in materiels_objs:
                        if m.est_disponible():
                            m.mettre_en_pret()
                elif statut == Emprunt.Statut.RETOURNE:
                    for m in materiels_objs:
                        if m.est_en_pret():
                            m.retourner()

                importes += 1

        except Exception as e:
            erreurs.append(f"Ligne {i} : erreur inattendue : {e}")
            ignores += 1
            continue

    if importes > 0 or mis_a_jour > 0:
        messages.success(
            request,
            f"Import terminé : {importes} créé(s), {mis_a_jour} mis à jour, {ignores} ignoré(s)."
        )
    else:
        messages.error(request, f"Aucun emprunt importé. {ignores} ligne(s) ignorée(s).")

    for erreur in erreurs[:10]:
        messages.warning(request, erreur)
    if len(erreurs) > 10:
        messages.warning(request, f"… et {len(erreurs) - 10} autre(s) erreur(s) non affichée(s).")

    return redirect(f"{reverse('equipement:dashboard')}?tab={TAB_RETOUR}")

def exporter_template_emprunts(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Emprunts"
    headers = ['id', 'materiel', 'emprunteur', 'date_emprunt', 'date_retour_prevue', 'notes', 'statut']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="16A34A")

    emprunts = Emprunt.objects.select_related(
        "materiels", "emprunteur"
    ).prefetch_related("lignes__materiel").order_by("id")

    for emprunt in emprunts:

        lignes = emprunt.lignes.all()
        if lignes.exists():
            noms_materiels = ", ".join(ligne.materiel.nom for ligne in lignes)
        else:
            noms_materiels = emprunt.materiel.nom if emprunt.materiel else ""

        ws.append([
            emprunt.id,
            noms_materiels,
            f"{emprunt.emprunteur.prenom} {emprunt.emprunteur.nom}",
            emprunt.date_operation.strftime("%Y-%m-%d %H:%M:%S") if emprunt.date_operation else "",
            emprunt.date_retour_prevue.strftime("%Y-%m-%d") if emprunt.date_retour_prevue else "",
            emprunt.notes or "",
            emprunt.statut or "",
        ])
    ws2 = wb.create_sheet(title="Materiels")
    ws2.append(['nom', 'id_materiel', 'etat'])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for m in Materiel.objects.all():
        ws2.append([m.nom, m.id_materiel, m.etat])
    ws3 = wb.create_sheet(title="Emprunteurs")
    ws3.append(['nom_complet', 'id_Tierce', 'type'])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    for e in Tierce.objects.all():
        ws3.append([f"{e.prenom} {e.nom}", e.id_Tierce, e.type_Tierce])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=emprunts.xlsx'
    wb.save(response)
    return response
